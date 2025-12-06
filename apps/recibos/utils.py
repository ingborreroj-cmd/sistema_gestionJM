import pandas as pd
from datetime import datetime
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging

# --- Django Imports ---
from django.db import transaction
from django.db.models import Q # <--- CORRECCIÓN 1: Importación de Q
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.text import slugify

# --- ReportLab Imports (Si está instalado) ---
# <--- CORRECCIÓN 2: Importaciones de ReportLab movidas al inicio
from reportlab.lib import pagesizes, colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter, landscape


# Importamos el modelo y las constantes definidas
from .models import Recibo, MAPEO_CATEGORIAS # Asume que MAPEO_CATEGORIAS existe

User = get_user_model()
logger = logging.getLogger(__name__) # <--- CORRECCIÓN 3: Configuración de Logging

# ----------------------------------------------------------------------
# --- 1. LÓGICA DE PROCESAMIENTO DE EXCEL (excel_procesor.py adaptado) ---
# ----------------------------------------------------------------------

def _obtener_valor_celda(row, col_name, default=None):
    """Obtiene un valor de una fila de DataFrame, manejando NaN y tipos."""
    valor = row.get(col_name)
    if pd.isna(valor) or valor is None or valor == "":
        return default
    
    # Manejo de números (RIF/Cédula, N° Recibo) para convertirlos a cadena limpia
    if col_name in ['RIF/Cédula', 'N° Recibo']:
        try:
            # Intenta convertir a int primero para eliminar decimales (.0) si es float
            return str(int(valor)).strip()
        except ValueError:
            # Si no se puede convertir a int (es texto o formato raro), retorna el string
            return str(valor).strip()
    
    return str(valor).strip()

@transaction.atomic
def procesar_excel_sync(file_content, user):
    """
    Procesa el contenido binario de un archivo Excel y crea/actualiza recibos.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_content))
    except Exception as e:
        logger.error(f"Error fatal al leer el archivo Excel: {e}")
        raise ValueError(f"Error al leer el archivo Excel: {e}") # Propaga error al usuario/vista

    filas_procesadas = 0
    recibos_creados = 0
    
    for index, row in df.iterrows():
        num_recibo = None
        try:
            # 1. Extracción de datos clave
            num_recibo = _obtener_valor_celda(row, 'N° Recibo')
            
            if not num_recibo:
                logger.warning(f"Fila {index}: Omitida, no tiene N° Recibo.")
                continue 

            # 2. Convertir Fecha
            fecha_excel = row.get('Fecha')
            fecha = None
            if isinstance(fecha_excel, datetime):
                fecha = fecha_excel.date()
            elif fecha_excel:
                # Intenta parsear como cadena (manejar múltiples formatos comunes)
                try:
                    fecha = datetime.strptime(str(fecha_excel).split()[0], '%Y-%m-%d').date()
                except ValueError:
                    # Alternativa común: D/M/Y
                    fecha = datetime.strptime(str(fecha_excel).split()[0], '%d/%m/%Y').date()
                except Exception as e:
                    logger.error(f"Fila {index} ({num_recibo}): Error al parsear fecha '{fecha_excel}': {e}")
                    raise 

            # 3. Crear o Actualizar Recibo
            recibo_data = {
                'estado': _obtener_valor_celda(row, 'Estado', default='Pagado'),
                'nombre': _obtener_valor_celda(row, 'Nombre'),
                'rif_cedula_identidad': _obtener_valor_celda(row, 'RIF/Cédula', default='N/A'),
                'direccion_inmueble': _obtener_valor_celda(row, 'Dirección'),
                'ente_liquidado': _obtener_valor_celda(row, 'Ente Liquidado'),
                'gastos_administrativos': Decimal(_obtener_valor_celda(row, 'Gastos Adm', default=0)),
                'tasa_dia': Decimal(_obtener_valor_celda(row, 'Tasa Día', default=1)),
                'total_monto_bs': Decimal(_obtener_valor_celda(row, 'Total Monto (Bs)', default=0)),
                'numero_transferencia': _obtener_valor_celda(row, 'N° Transferencia'),
                'conciliado': (_obtener_valor_celda(row, 'Conciliado') in ['Sí', 'SI', 'si', True, 1]),
                'fecha': fecha,
                'concepto': _obtener_valor_celda(row, 'Concepto', default='N/A'),
            }
            
            # 4. Asignar Categorías (Booleano)
            for i in range(1, 11):
                col_name = f'Categoría {i}'
                field_name = f'categoria{i}'
                valor = _obtener_valor_celda(row, col_name)
                recibo_data[field_name] = (valor in ['Sí', 'SI', 'si', True, 1])

            # 5. Obtener o crear el recibo
            recibo, created = Recibo.objects.update_or_create(
                numero_recibo=num_recibo,
                defaults={**recibo_data, 'usuario_creador': user}
            )
            
            if created:
                recibos_creados += 1
            
            filas_procesadas += 1

        except Exception as e:
            logger.error(f"Error procesando fila {index} (N° Recibo: {num_recibo}): {e}")
            # La transacción no se revierte aquí, solo se omite la fila defectuosa.
            # Si quieres que toda la carga falle si una fila falla, comenta el 'continue'
            continue 

    return {'total_procesados': filas_procesadas, 'nuevos_recibos': recibos_creados}

# ---------------------------------------------------------------------
# --- 2. LÓGICA DE FILTRADO (report_generator.py - Parte 1) ---
# ---------------------------------------------------------------------

def obtener_recibos_filtrados(fecha_inicio, fecha_fin, estado_filtro, categorias_filtro):
    """
    Construye el queryset de recibos basándose en los criterios del formulario de reportes.
    """
    queryset = Recibo.objects.all()

    # Filtro de Rango de Fechas
    if fecha_inicio:
        queryset = queryset.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha__lte=fecha_fin)
    
    # Filtro por Estado 
    if estado_filtro and estado_filtro != 'Todos':
        if estado_filtro.lower() == 'anulado':
             queryset = queryset.filter(anulado=True)
        elif estado_filtro.lower() == 'activo':
             queryset = queryset.filter(anulado=False)
        else:
             queryset = queryset.filter(estado__iexact=estado_filtro)
    
    # Filtro por Categorías (Lógica OR)
    if categorias_filtro:
        q_objects = Q()
        for cat_index in categorias_filtro:
            field_name = f'categoria{cat_index}'
            # Acumula condiciones de OR (ej: (categoria1=True) OR (categoria3=True))
            q_objects |= Q(**{field_name: True}) 
        queryset = queryset.filter(q_objects)
        
    return queryset.order_by('-fecha', 'numero_recibo')

# ---------------------------------------------------------------------
# --- 3. LÓGICA DE GENERACIÓN DE REPORTES EXCEL
# ---------------------------------------------------------------------

def generar_reporte_excel_django(queryset, filtros_data, user_name):
    """
    Genera el reporte Excel con los datos del queryset y lo devuelve como HttpResponse.
    """
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    hoja_datos = workbook.active
    hoja_datos.title = "Recibos_Filtrados"

    columns = [
        'N° Recibo', 'Fecha', 'Estado', 'Nombre', 'Cédula/RIF', 'Monto Total (Bs)',
        'Conciliado', 'N° Transf.', 'Gasto Adm.', 'Tasa Día', 'Usuario Creador', 
        'Categorías Seleccionadas'
    ]
    hoja_datos.append(columns)
    
    # Estilos de encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col_idx, column in enumerate(columns, 1):
        cell = hoja_datos.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill

    # Llenar datos
    for recibo in queryset:
        # Asume que get_categorias_marcadas() retorna una cadena de texto (ej: "Cat 1, Cat 5")
        categorias_marcadas = recibo.get_categorias_marcadas() 
        
        hoja_datos.append([
            recibo.numero_recibo,
            recibo.fecha.strftime('%Y-%m-%d'),
            "ANULADO" if recibo.anulado else recibo.estado,
            recibo.nombre,
            recibo.rif_cedula_identidad,
            float(recibo.total_monto_bs),
            "Sí" if recibo.conciliado else "No",
            recibo.numero_transferencia,
            float(recibo.gastos_administrativos),
            float(recibo.tasa_dia),
            recibo.usuario_creador.username,
            categorias_marcadas,
        ])
        
    # Guardar el libro de trabajo
    workbook.save(output)
    output.seek(0)

    # Devolver la respuesta HTTP
    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Reporte_Recibos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

# ---------------------------------------------------------------------
# --- 4. LÓGICA DE GENERACIÓN DE PDF INDIVIDUAL
# ---------------------------------------------------------------------

def generar_pdf_individual(recibo):
    """
    Genera el PDF individual de un recibo usando ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    Story = []

    # Título
    Story.append(Paragraph(f"**RECIBO DE PAGO N° {recibo.numero_recibo}**", styles['h1']))
    Story.append(Spacer(1, 0.2 * inch))
    
    # Estado (destacado si está anulado)
    if recibo.anulado:
        estado_texto = f"⚠️ ESTADO: **ANULADO** ⚠️"
        Story.append(Paragraph(estado_texto, styles['h2']))
        Story.append(Spacer(1, 0.1 * inch))


    # Detalles del Cliente
    Story.append(Paragraph(f"**Cliente:** {recibo.nombre}", styles['Normal']))
    Story.append(Paragraph(f"**Cédula/RIF:** {recibo.rif_cedula_identidad}", styles['Normal']))
    Story.append(Paragraph(f"**Dirección:** {recibo.direccion_inmueble or 'N/A'}", styles['Normal']))
    Story.append(Paragraph(f"**Fecha de Emisión/Pago:** {recibo.fecha.strftime('%d/%m/%Y')}", styles['Normal']))
    Story.append(Spacer(1, 0.1 * inch))

    # Tabla de Montos
    data = [
        ['Concepto', 'Monto (Bs.)'],
        ['Total Recibo', f"Bs. {recibo.total_monto_bs:,.2f}"],
        ['Gastos Administrativos', f"Bs. {recibo.gastos_administrativos:,.2f}"],
        ['Tasa del Día', f"{recibo.tasa_dia:,.4f}"],
        ['N° Transferencia', recibo.numero_transferencia or 'N/A'],
        ['Conciliado', "Sí" if recibo.conciliado else "No"],
    ]
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#004D99')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ])

    t = Table(data, colWidths=[3*inch, 2*inch])
    t.setStyle(table_style)
    Story.append(t)
    Story.append(Spacer(1, 0.5 * inch))
    
    # Pie de página y Auditoría
    Story.append(Paragraph(f"**Concepto:** {recibo.concepto}", styles['Normal']))
    Story.append(Spacer(1, 0.2 * inch))
    Story.append(Paragraph(f"**Creado por:** {recibo.usuario_creador.username} el {recibo.fecha_creacion.strftime('%d/%m/%Y')}", styles['Italic']))

    doc.build(Story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"Recibo_{recibo.numero_recibo}.pdf"
    response['Content-Disposition'] = f'inline; filename={filename}'
    return response

# ---------------------------------------------------------------------
# --- 5. LÓGICA DE GENERACIÓN DE REPORTE PDF COMPLETO
# ---------------------------------------------------------------------

def generar_reporte_pdf_django(queryset, filtros_data, user_name):
    """
    Genera el reporte PDF completo de la lista de recibos filtrados.
    """
    buffer = io.BytesIO()
    # Usar landscape para una tabla ancha de reporte
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                             topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    Story = []
    
    # Título y Filtros
    Story.append(Paragraph(f"**REPORTE CONSOLIDADO DE RECIBOS**", styles['h1']))
    Story.append(Paragraph(f"Generado por: {user_name} el {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    # Aquí podrías añadir los detalles de 'filtros_data'
    Story.append(Spacer(1, 0.2 * inch))

    # Data para la tabla principal
    table_data = []
    headers = [
        'N° Recibo', 'Fecha', 'Estado', 'Nombre', 'Monto (Bs)', 'Conciliado', 'Categorías'
    ]
    table_data.append(headers)

    for recibo in queryset:
        categorias_marcadas = recibo.get_categorias_marcadas() 
        table_data.append([
            recibo.numero_recibo,
            recibo.fecha.strftime('%d/%m/%Y'),
            "ANULADO" if recibo.anulado else recibo.estado,
            recibo.nombre,
            f"Bs. {recibo.total_monto_bs:,.2f}",
            "Sí" if recibo.conciliado else "No",
            categorias_marcadas,
        ])
    
    # Estilo de la tabla
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ])
    
    # Calcula el ancho de las columnas (el ancho total es 10.5 pulgadas en landscape letter)
    col_widths = [1*inch, 1*inch, 1*inch, 3*inch, 1.5*inch, 1*inch, 2*inch] 
    
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(table_style)
    Story.append(t)

    doc.build(Story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"Reporte_Consolidado_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}' 
    return response